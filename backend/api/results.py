"""
Results endpoints — list scores, get detail, delete, export
"""
from fastapi import APIRouter, HTTPException, Header, Query
from fastapi.responses import StreamingResponse
from typing import Optional
import io
from db.supabase_client import supabase

router = APIRouter(prefix="/api/v1/results", tags=["results"])


def _get_user(authorization: Optional[str]) -> str:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Not authenticated")
    token = authorization.replace("Bearer ", "")
    try:
        user = supabase.auth.get_user(token)
        return user.user.id
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")


@router.get("")
async def list_results(
    job_id: Optional[str] = Query(None),
    recommendation: Optional[str] = Query(None),
    authorization: Optional[str] = Header(None),
):
    _get_user(authorization)
    try:
        query = supabase.table("scores")\
            .select("*, candidates(name, email, phone)")\
            .order("overall_score", desc=True)

        if job_id:
            query = query.eq("job_id", job_id)
        if recommendation:
            query = query.eq("recommendation", recommendation.upper())

        res = query.execute()
        return {"results": res.data, "total": len(res.data)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/{score_id}")
async def get_result(
    score_id: str,
    authorization: Optional[str] = Header(None),
):
    _get_user(authorization)
    try:
        res = supabase.table("scores")\
            .select("*, candidates(name, email, phone)")\
            .eq("id", score_id)\
            .execute()
        if not res.data:
            raise HTTPException(status_code=404, detail="Score not found")
        return {"result": res.data[0]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/{score_id}")
async def delete_result(
    score_id: str,
    authorization: Optional[str] = Header(None),
):
    _get_user(authorization)
    try:
        supabase.table("scores").delete().eq("id", score_id).execute()
        return {"message": "Deleted"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export/excel")
async def export_excel(
    job_id: Optional[str] = Query(None),
    shortlisted_only: bool = Query(False),
    authorization: Optional[str] = Header(None),
):
    _get_user(authorization)
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

        query = supabase.table("scores")\
            .select("*, candidates(name, email, phone)")\
            .order("overall_score", desc=True)

        if job_id:
            query = query.eq("job_id", job_id)
        if shortlisted_only:
            query = query.eq("recommendation", "SHORTLIST")

        res = query.execute()
        scores = res.data

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ATS Results"

        # Header style
        header_fill = PatternFill("solid", fgColor="1B4F8A")
        header_font = Font(color="FFFFFF", bold=True)

        headers = [
            "Rank", "Name", "Email", "Phone", "Overall Score",
            "Status", "Technical", "Experience", "Education",
            "Soft Skills", "Stability", "Matched Skills",
            "Missing Skills", "AI Reasoning", "Model Used"
        ]

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Color fills
        green  = PatternFill("solid", fgColor="D4EDDA")
        yellow = PatternFill("solid", fgColor="FFF3CD")
        red    = PatternFill("solid", fgColor="F8D7DA")

        for ri, score in enumerate(scores, 2):
            candidate = score.get("candidates") or {}
            rec = score.get("recommendation", "PASS")
            fill = green if rec == "SHORTLIST" else yellow if rec == "REVIEW" else red
            cats = score.get("category_scores", {})

            row = [
                ri - 1,
                candidate.get("name", "Unknown"),
                candidate.get("email", ""),
                candidate.get("phone", ""),
                f"{score.get('overall_score', 0)}%",
                rec,
                f"{cats.get('technical', {}).get('score', 0)}%",
                f"{cats.get('experience', {}).get('score', 0)}%",
                f"{cats.get('education', {}).get('score', 0)}%",
                f"{cats.get('soft_skills', {}).get('score', 0)}%",
                f"{cats.get('stability', {}).get('score', 0)}%",
                ", ".join(score.get("matched_skills", [])),
                ", ".join(score.get("missing_skills", [])),
                score.get("ai_reasoning", ""),
                score.get("model_used", ""),
            ]

            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if ci in (5, 6):
                    cell.fill = fill

        # Column widths
        widths = [6,20,25,15,14,12,12,12,12,12,12,35,30,50,12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(i)
            ].width = w

        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=ATS_Results.xlsx"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
